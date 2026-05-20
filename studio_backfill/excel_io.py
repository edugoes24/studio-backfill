"""Read the studio_results Excel and write the augmented output Excel."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Iterator

import openpyxl


# The columns we care about (case-sensitive header names in the source Excel).
REQUIRED_COLUMNS = [
    "id", "instrumentId", "tutorId", "teacherId", "sectionId",
    "schoolCode", "payload", "submittedAt", "utilitiesLink",
]


def read_rows(excel_path: str) -> Iterator[dict]:
    """Yield each data row of the first sheet as a dict keyed by header.

    Coerces datetimes to ISO 8601 strings and leaves JSON-string fields alone.
    The caller can parse `utilitiesLink` and `payload` from JSON as needed.
    """
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows_iter = ws.iter_rows(values_only=True)
    header = list(next(rows_iter))
    missing = [c for c in REQUIRED_COLUMNS if c not in header]
    if missing:
        raise RuntimeError(f"Excel missing required columns: {missing}")

    for raw in rows_iter:
        if raw is None:
            continue
        row = dict(zip(header, raw))
        if row.get("id") is None:
            continue
        if hasattr(row.get("submittedAt"), "isoformat"):
            row["submittedAt"] = row["submittedAt"].isoformat()
        yield row


def parse_utilities(row: dict) -> dict:
    raw = row.get("utilitiesLink")
    if not raw:
        return {}
    if isinstance(raw, dict):
        return raw
    try:
        return json.loads(raw)
    except (TypeError, ValueError):
        return {}


def parse_payload(row: dict) -> dict:
    raw = row.get("payload")
    if not raw:
        return {}
    if isinstance(raw, dict):
        return raw
    try:
        return json.loads(raw)
    except (TypeError, ValueError):
        return {}


def write_output_excel(
    source_excel_path: str,
    output_path: str,
    rows_with_links: dict[int, dict],
    source_sha256: str | None = None,
) -> None:
    """Generate output Excel: source columns + event_id_xai + pdf_drive_link + backfill_status.

    `rows_with_links` maps excel_id → {event_id_xai, pdf_drive_link, backfill_status}.
    """
    src = Path(source_excel_path)
    if not src.exists():
        raise FileNotFoundError(source_excel_path)

    wb = openpyxl.load_workbook(source_excel_path)
    ws = wb[wb.sheetnames[0]]

    # Add the 3 new columns at the right
    header = [cell.value for cell in ws[1]]
    new_cols = ["event_id_xai", "pdf_drive_link", "backfill_status"]
    start_col = len(header) + 1
    for i, col_name in enumerate(new_cols):
        ws.cell(row=1, column=start_col + i, value=col_name)

    id_idx = header.index("id") + 1
    for r in range(2, ws.max_row + 1):
        excel_id = ws.cell(row=r, column=id_idx).value
        if excel_id is None:
            continue
        info = rows_with_links.get(int(excel_id), {})
        ws.cell(row=r, column=start_col + 0, value=info.get("event_id_xai", ""))
        ws.cell(row=r, column=start_col + 1, value=info.get("pdf_drive_link", ""))
        ws.cell(row=r, column=start_col + 2, value=info.get("backfill_status", ""))

    # SHA-256 trazability cell: write as a comment-like text in a new "meta" sheet
    # to avoid touching the data sheet structure further.
    if source_sha256:
        meta_sheet = wb.create_sheet("_backfill_meta")
        meta_sheet["A1"] = "source_excel_path"
        meta_sheet["B1"] = source_excel_path
        meta_sheet["A2"] = "source_excel_sha256"
        meta_sheet["B2"] = source_sha256
        meta_sheet["A3"] = "generated_by"
        meta_sheet["B3"] = "studio_backfill"

    wb.save(output_path)
